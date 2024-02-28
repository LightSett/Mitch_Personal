import openpyxl
import os

def combine_and_filter_excel_files(folder_path):
    combined_wb = openpyxl.Workbook()  # Create a new workbook for combined data
    combined_sheet = combined_wb.active
    combined_sheet.title = 'CombinedData'  # Set a sheet name for combined data

    header_copied = False  # Flag to copy the header only once

    # Initialize variable to keep track of row index in combined sheet
    combined_row = 1

    # Iterate through all .xlsx files in the folder
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
            file_path = os.path.join(folder_path, file_name)
            try:
                # Load the workbook
                wb = openpyxl.load_workbook(file_path)

                # Flag to check if the header is already copied
                header_copied_in_current_file = False

                # Iterate through each sheet in the source file
                for sheet in wb.sheetnames:
                    source_sheet = wb[sheet]

                    # Iterate through rows in the source sheet
                    for row_num, row in enumerate(source_sheet.iter_rows(), start=1):
                        if not header_copied and not header_copied_in_current_file:
                            # Copy the header from the first sheet
                            values = [cell.value for cell in row]
                            combined_sheet.append(values)
                            combined_row += 1
                            header_copied = True
                            header_copied_in_current_file = True
                        elif row_num != 1 or (row[5].value and "pole" in str(row[5].value).lower()):
                            values = [cell.value for cell in row]
                            combined_sheet.append(values)
                            combined_row += 1

            except Exception as e:
                print(f"Error processing file {file_path}: {str(e)}")

    # Filter rows in the combined sheet based on column F (except for the first row)
    for row in combined_sheet.iter_rows(min_row=2, min_col=6, max_col=6):
        if not (row[0].value and "pole" in str(row[0].value).lower()):
            for cell in row:
                cell.value = None

    # Save the filtered combined data to a new Excel file
    combined_file_path = os.path.join(folder_path, "MR_Combined_XLSX.xlsx")
    combined_wb.save(combined_file_path)
    print(f"Filtered combined Excel file saved at: {combined_file_path}")

# Replace 'folder_path' with the actual folder path containing your Excel files
folder_path = r"C:\Users\mmccarthy\Desktop\1101PA Make Ready Info\XLSX"
combine_and_filter_excel_files(folder_path)
