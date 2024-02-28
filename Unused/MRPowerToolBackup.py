import openpyxl
import os
import csv
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import threading
import tempfile

# Variable to indicate if the process is running
running = False

def combine_and_filter_excel_files(folder_path):
    global running
    running = True

    print("Starting the combine and filter process...")

    try:
        combined_wb = openpyxl.Workbook()  # Create a new workbook for combined data
        combined_sheet = combined_wb.active
        combined_sheet.title = 'CombinedData'  # Set a sheet name for combined data

        header_copied = False  # Flag to copy the header only once

        # Initialize variable to keep track of row index in combined sheet
        combined_row = 1

        # Iterate through all .xlsx files in the folder
        for file_name in os.listdir(folder_path):
            if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
                file_path = os.path.join(folder_path, file_name)
                try:
                    # Load the workbook
                    wb = openpyxl.load_workbook(file_path)

                    # Flag to check if the header is already copied
                    header_copied_in_current_file = False

                    # Get the jobname from the filename
                    filename_numerals = ''.join(filter(str.isdigit, file_name))  # Extracting numerals from filename

                    # Iterate through each sheet in the source file
                    for sheet in wb.sheetnames:
                        source_sheet = wb[sheet]

                        # Check if column D ('jobname') is already populated
                        if source_sheet.cell(row=1, column=4).value != "jobname":
                            source_sheet.cell(row=1, column=4, value="jobname")  # Setting cell D1 as "jobname"

                            # Start populating 'jobname' from the second row (row index 2)
                            for row_num, row in enumerate(source_sheet.iter_rows(min_row=2), start=2):
                                # Calculate the 'jobname' value based on the filename numerals
                                jobname_value = f"APP_{filename_numerals}"

                                # Set the 'jobname' value in column D for each row
                                source_sheet.cell(row=row_num, column=4, value=jobname_value)

                        # Iterate through rows in the source sheet
                        for row_num, row in enumerate(source_sheet.iter_rows(), start=1):
                            if not header_copied and not header_copied_in_current_file:
                                # Copy the header from the first sheet
                                values = [cell.value for cell in row]
                                combined_sheet.append(values)
                                combined_row += 1
                                header_copied = True
                                header_copied_in_current_file = True

                            elif row_num != 1 or (row[5].value and "pole" in str(row[5].value).lower()):
                                values = [cell.value for cell in row]
                                combined_sheet.append(values)
                                combined_row += 1

                except Exception as e:
                    print(f"Error processing file {file_path}: {str(e)}")

       # Filter rows in the combined sheet based on column F (except for the first row)
        rows_to_delete = []
        for row_index, row in enumerate(combined_sheet.iter_rows(min_row=2, min_col=6, max_col=6), start=2):
            if not (row[0].value and "pole" in str(row[0].value).lower()):
                rows_to_delete.append(row_index)

        # Delete the rows that do not meet the criteria
        for row_index in reversed(rows_to_delete):
            combined_sheet.delete_rows(row_index)
        
        # Save the changes to the workbook
        combined_wb.save(os.path.join(folder_path, "CombinedData.xlsx"))

        for file_name in os.listdir(folder_path):
            if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
                try:
                    xlsx_file_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.load_workbook(xlsx_file_path)
                    sheet = wb.active

                    csv_file_path = os.path.join(folder_path, file_name.replace('.xlsx', '.csv'))
                    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csv_file:
                        writer = csv.writer(csv_file)
                        for row in sheet.iter_rows():
                            writer.writerow([cell.value for cell in row])
                        
                    # Rename the CSV file to remove 'XLSX' or '_XLSX' from the name
                    new_csv_file_path = csv_file_path.replace(' XLSX', '_CSV').replace('_XLSX', '_CSV')  # Remove 'XLSX' or '_XLSX' from the filename
                    os.rename(csv_file_path, new_csv_file_path)
                except Exception as e:
                    print(f"Error converting {file_name} to CSV: {str(e)}")
               
        print("Combine and filter process completed successfully.")
    except Exception as e:
        print(f"Error during the combine and filter process: {str(e)}")
    running = False




def remove_invalid_rows(folder_path, entry=None):
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.csv') and not file_name.startswith('~$'):
            csv_file_path = os.path.join(folder_path, file_name)
            temp_file_path = os.path.join(folder_path, "temp.csv")  # Temporary file to store valid rows

            with open(csv_file_path, 'r', newline='', encoding='utf-8') as csvfile, open(temp_file_path, 'w', newline='', encoding='utf-8') as temp:
                reader = csv.reader(csvfile)
                writer = csv.writer(temp)

                for row in reader:
                    # Checking if column F is empty or doesn't contain "pole"
                    if len(row) >= 6 and (not row[5] or row[5].strip().lower() != "pole"):
                        continue  # Skip invalid rows

                    writer.writerow(row)

            os.remove(csv_file_path)  # Remove the original CSV file
            os.rename(temp_file_path, csv_file_path)  # Rename the temp file to the original file name
    
    if entry:
        # Show popup message when the script is finished
        messagebox.showinfo("Process Completed", "Invalid rows removed from CSV files.")
        # Reset GUI elements here
        entry.delete(0, tk.END)


# Function to populate a new column with 'jobname' in CSV files
def populate_new_column_with_jobname(folder_path, entry=None):
    combined_data = []  # To store combined data

    for file_name in os.listdir(folder_path):
        if file_name.endswith('.csv') and not file_name.startswith('~$'):
            file_path = os.path.join(folder_path, file_name)
            try:
                # Read the CSV file
                with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    data = list(reader)

                # Update data in a new column with 'jobname'
                header = data[0]  # Extracting header
                if header[0] != 'jobname':  # Check if 'jobname' header is already present
                    header.insert(0, 'jobname')  # Adding 'jobname' to the header

                for idx, row in enumerate(data[1:], start=1):
                    filename_parts = file_name.split(' ')
                    jobname_parts = [part for part in filename_parts if not part.isnumeric()]  # Extract non-numeric parts
                    jobname = '_'.join(jobname_parts)  # Joining the non-numeric parts
                    jobname = ''.join(filter(str.isdigit, jobname))  # Remove everything after numerals
                    row.insert(0, jobname)  # Insert 'jobname' as a new column value

                    # Add "APP_" to the contents of the populated column A except for the header
                    if len(row) > 0 and idx != 0:
                        row[0] = 'APP_' + row[0] if row[0] else ''  # Add 'APP_' to non-empty rows

                # Write the updated data back to the CSV file
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerows(data)

                # Append data to combined_data
                combined_data.extend(data[1:])  # Exclude the header

            except Exception as e:
                print(f"Error processing file {file_name}: {str(e)}")

    # Write combined data to a new CSV file
    combined_file_path = os.path.join(folder_path, "MRSummary_Combined.csv")
    with open(combined_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)  # Write header for combined file
        writer.writerows(combined_data)

    for file_name in os.listdir(folder_path):
            if file_name.endswith('.csv') and not file_name.startswith('~$'):
                try:
                    csv_file_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.Workbook()
                    sheet = wb.active

                    with open(csv_file_path, 'r', newline='', encoding='utf-8') as csv_file:
                        reader = csv.reader(csv_file)
                        for row in reader:
                            sheet.append(row)

                    # Create an XLSX file name based on the original CSV file name
                    xlsx_file_path = os.path.join(folder_path, f"{file_name.split('.')[0]}.xlsx")

                    # Save the workbook as an XLSX file
                    wb.save(xlsx_file_path)
                    print(f"Converted {file_name} to XLSX.")

                except Exception as e:
                    print(f"Error converting {file_name} to XLSX: {str(e)}")

    if entry:
        # Show popup message when the script is finished
        messagebox.showinfo("Process Completed", "New column 'jobname' added with values. Combined CSV created.")
        # Reset GUI elements here
        entry.delete(0, tk.END)

def get_directory():
    def select_folder():
        folder_path = filedialog.askdirectory()
        if folder_path:
            entry.delete(0, tk.END)
            entry.insert(0, folder_path)

    def run_process():
        folder_path = entry.get()
        if folder_path and not running:
            thread = threading.Thread(target=combine_and_filter_excel_files, args=(folder_path,))
            thread.start()

    def run_csv_process():
        folder_path = entry.get()
        if folder_path and not running:
            thread = threading.Thread(target=remove_invalid_rows, args=(folder_path, entry))
            thread.start()

    def run_jobname_population():
        folder_path = entry.get()
        if folder_path and not running:
            thread = threading.Thread(target=populate_new_column_with_jobname, args=(folder_path, entry))
            thread.start()


    def close_script():
        global running
        if running:
            tk.messagebox.showinfo("Script Running", "Script Running - Please Wait")
        else:
            root.destroy()

    root = tk.Tk()
    root.title("MRPowerTool")
    root.geometry("400x650")

    label = tk.Label(root, text="The Make Ready Power Tool:", font=("Arial Bold", 14))
    label.pack(pady=5)

    note_text = "Welcome to the Make Ready Powertool.\n \nFollow the steps below to generate the necessary files for Make Ready.\n \n \nStep 1:\nPaste or Navigate to the folder in the MR Data containing only XLSX files.\nIf no folder exists, create a folder in your MR Data that contains only "
    note_text += '"APP_(App Number) Node Attributes with IDs XLSX."'
    note = tk.Label(root, text=note_text, font=("Arial", 10), justify="center", wraplength=380)
    note.pack(pady=5)

    entry = tk.Entry(root, font=("Arial", 10), width=40)
    entry.pack(pady=5)

    browse_button = tk.Button(root, text="Browse", command=select_folder)
    browse_button.pack(pady=5)

    run_button = tk.Button(root, text="Create and Combine CSVs", command=run_process)
    run_button.pack(pady=5)

    #run_csv_button = tk.Button(root, text="Clean Up CSVs", command=run_csv_process)
    #run_csv_button.pack(pady=5)

    note_text = "\nStep 2:\n Using the same browse window above or pasting a new filepath above, navigate to your MakeReadySummary Folder containing only CSV files.\nIf no folder exists, create a folder in your MR Data that contains only "
    note_text += '"APP_(App Number) Make Ready Summary"'
    note = tk.Label(root, text=note_text, font=("Arial", 10), justify="center", wraplength=380)
    note.pack(pady=5)

    populate_button = tk.Button(root, text="Assign Summary Jobname, Convert, & Combine", command=lambda: populate_new_column_with_jobname(entry.get(), entry))
    populate_button.pack(pady=5)

    note_text = "\nCheck to see that your folders contain both individual and combined XLSX and CSV Files.\nFollow the FlexNAP Make Ready Guide for the next steps."
    note = tk.Label(root, text=note_text, font=("Arial", 10), justify="center", wraplength=380)
    note.pack(pady=5)

    note_text = "Good Luck!"
    note = tk.Label(root, text=note_text, font=("Arial Bold", 10), justify="center", wraplength=380)
    note.pack(pady=5)   

    close_button = tk.Button(root, text="Close", command=close_script)
    close_button.pack(pady=5)

    root.mainloop()

get_directory()
