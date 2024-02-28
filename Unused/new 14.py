import openpyxl

def update_excel_jobname(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the active sheet
    sheet = wb.active
    
    # Update header for column D
    sheet['D1'] = 'jobname'
    
    # Iterate through rows in column A (assuming the titles are in column A)
    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
        title = row[0].value
        
        # Extract characters before the first space
        if title:
            job_name = title.split(' ')[0]
            
            # Write the extracted data to column D
            sheet.cell(row=row[0].row, column=4).value = job_name
    
    # Save the changes to the Excel file
    wb.save(file_path)
    print(f"Job names added to column D in {file_path}")

# Replace 'your_file.xlsx' with the actual path to your Excel file
excel_file_path = r'C:\Users\mmccarthy\Desktop\1101PA Make Ready Info\XLSX\APP_306524 Node Attributes with IDs XLSX.xlsx'
update_excel_jobname(excel_file_path)
