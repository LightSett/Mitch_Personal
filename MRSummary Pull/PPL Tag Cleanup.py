import os
import re
from openpyxl import load_workbook

# Function to clean up PPL Tag IDs
def clean_ppl_tag_id(tag):
    pattern = r'\b\d{0,5}N\d{0,5}\b'  # Regex pattern for PPL Tag ID with word boundary
    matches = re.findall(pattern, tag)
    return matches[0] if matches else None

# Function to remove rows where column A is blank
def remove_blank_rows(ws):
    rows_to_remove = []
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        if not row[0].value:  # Check if column A is empty
            rows_to_remove.append(row)

    for row in rows_to_remove:
        ws.delete_rows(row[0].row)

# Function to clean up column C in Excel file
def clean_column_c(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        remove_blank_rows(ws)  # Remove rows where column A is blank

        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            for cell in row:
                if cell.column == 3:  # Assuming column C is the third column
                    tag_id = clean_ppl_tag_id(str(cell.value))
                    if tag_id:
                        cell.value = tag_id

    wb.save(file_path)

# Directory containing .xlsx files
directory = r'C:\Users\mmccarthy\Desktop\Automation Team\Lightsett\MRSummary Pull\MakeReadySummaryAttributes'

# Iterate through each file in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)
        clean_column_c(file_path)
