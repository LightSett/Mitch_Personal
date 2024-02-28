import arcpy
import os
from openpyxl import Workbook, load_workbook
import string

# Function to export MakeReadySummary attribute table to Excel
def export_make_ready_summary(gdb_path, output_folder):
    arcpy.env.workspace = gdb_path
    make_ready_summary = "MakeReadySummary"

    if arcpy.Exists(make_ready_summary):
        # Get the name of the parent folder containing the GDB
        parent_folder_name = os.path.basename(os.path.dirname(gdb_path))

        # Cleanse file name to remove invalid characters
        valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
        sanitized_name = ''.join(c for c in parent_folder_name if c in valid_chars)
        output_excel = os.path.join(output_folder, f"{sanitized_name}_{make_ready_summary}_Attributes.xlsx")
        arcpy.TableToExcel_conversion(make_ready_summary, output_excel)
    else:
        print(f"No {make_ready_summary} feature found in {gdb_path}")

def read_gdbs_from_excel(xlsx_file):
    gdbs = []

    try:
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            gdb_path = row[0]  # Assuming GDB path is in the first column
            gdbs.append(gdb_path)
    except Exception as e:
        print(f"Error reading the Excel file: {e}")

    return gdbs

def main():
    # Paths and file names
    xlsx_file = r"C:\Users\mmccarthy\Desktop\Automation Team\Lightsett\MRSummary Pull\LLD_LLD_List.xlsx"
    output_directory = os.path.join(os.path.dirname(__file__), "MakeReadySummaryAttributes")

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Read gdbs from xlsx file
    gdbs = read_gdbs_from_excel(xlsx_file)

    if not gdbs:
        print("No GDB paths found in the input Excel file.")
        return

    # Process each GDB
    for gdb_path in gdbs:
        export_make_ready_summary(gdb_path, output_directory)

if __name__ == "__main__":
    main()
