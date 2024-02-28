import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import tkinter as tk
from tkinter import filedialog, messagebox

selected_directory = ""

def process_excel_files(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(directory, filename)
            wb = load_workbook(filepath)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                merge_text_bold(ws, 2, "Splice Closures and Equipment")
                merge_text_bold(ws, 73, "Fiber Optic Cables")
                merge_text_bold(ws, 118, "Civil Items")
            wb.save(filepath)
    messagebox.showinfo("Success", "Operations completed on Excel files!")

def merge_text_bold(ws, row_num, text):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1)
    cell.value = text
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def select_directory():
    global selected_directory
    directory = filedialog.askdirectory()
    if directory:
        selected_directory = directory
        directory_label.config(text="Selected Directory: " + selected_directory)

def clear_contents_of_specified_rows():
    if selected_directory:
        for filename in os.listdir(selected_directory):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(selected_directory, filename)
                wb = load_workbook(filepath)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    merge_text_bold(ws, 2, "Splice Closures and Equipment")
                    merge_text_bold(ws, 73, "Fiber Optic Cables")
                    merge_text_bold(ws, 118, "Civil Items")
                wb.save(filepath)
        messagebox.showinfo("Success", "Merging completed on Excel files!")
    else:
        messagebox.showwarning("Warning", "Please select a directory!")

def initialize_script():
    if selected_directory:
        process_excel_files(selected_directory)
        clear_contents_of_specified_rows()
    else:
        messagebox.showwarning("Warning", "Please select a directory!")

root = tk.Tk()
root.title("Excel File Processing")

directory_label = tk.Label(root, text="Selected Directory: " + selected_directory)
directory_label.pack(padx=10, pady=5)

select_button = tk.Button(root, text="Select Directory", command=select_directory)
select_button.pack(padx=10, pady=5)

run_button = tk.Button(root, text="Run", command=initialize_script)
run_button.pack(padx=10, pady=5)

close_button = tk.Button(root, text="Close", command=root.destroy)
close_button.pack(padx=10, pady=5)

root.mainloop()
