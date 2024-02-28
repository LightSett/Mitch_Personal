import os
from tkinter import filedialog, messagebox, Tk, Label, Entry, Button
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def modify_excel(directory):
    excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]

    for file_name in excel_files:
        file_path = os.path.join(directory, file_name)

        try:
            wb = load_workbook(file_path)
            sheet = wb.active

            black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            color_B7C9E2 = PatternFill(start_color='B7C9E2', end_color='B7C9E2', fill_type='solid')
            color_d8d8d8 = PatternFill(start_color='D8D8D8', end_color='D8D8D8', fill_type='solid')

            # Change cell fill color to black for rows 72 and 117 in columns A to F
            for row_number in [72, 117]:
                for col in range(1, 7):  # Columns A to F
                    sheet.cell(row=row_number, column=col).fill = black_fill

            # Clear contents of rows 139 and 140 in columns A to F
            for row_number in [139, 140]:
                for col in range(1, 7):  # Columns A to F
                    sheet.cell(row=row_number, column=col).value = None

            # Change cell fill color to B7C9E2 for rows 2, 73, and 118 in columns A to F
            for row_number in [2, 73, 118]:
                for col in range(1, 7):  # Columns A to F
                    sheet.cell(row=row_number, column=col).fill = color_B7C9E2

            # Change cell fill color to d8d8d8 for specified rows and columns
            for row_number in range(4, 71):  # Even rows 4-70
                if row_number % 2 == 0:
                    for col in range(1, 7):  # Columns A to F
                        sheet.cell(row=row_number, column=col).fill = color_d8d8d8

            for row_number in range(75, 116):  # Odd rows 75-115
                if row_number % 2 != 0:
                    for col in range(1, 7):  # Columns A to F
                        sheet.cell(row=row_number, column=col).fill = color_d8d8d8

            for row_number in range(120, 139):  # Even rows 120-138
                if row_number % 2 == 0:
                    for col in range(1, 7):  # Columns A to F
                        sheet.cell(row=row_number, column=col).fill = color_d8d8d8

            wb.save(file_path)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error processing {file_name}: {str(e)}")

def browse_directory():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        directory_entry.delete(0, 'end')
        directory_entry.insert('end', folder_selected)

def run_operation():
    directory = directory_entry.get()
    if directory:
        modify_excel(directory)
        messagebox.showinfo("Operation Complete", "Modifications applied to Excel files in the selected directory.")
    else:
        messagebox.showwarning("Warning", "Please select a directory.")

root = Tk()
root.title("Excel Modifications")

# Create GUI components
label = Label(root, text="Select a directory:")
label.pack()

directory_entry = Entry(root, width=50)
directory_entry.pack()

browse_button = Button(root, text="Browse", command=browse_directory)
browse_button.pack()

run_button = Button(root, text="Run", command=run_operation)
run_button.pack()

close_button = Button(root, text="Close", command=root.destroy)
close_button.pack()

root.mainloop()
