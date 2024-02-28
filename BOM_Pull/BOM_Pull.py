import arcpy
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # Add this import line for get_column_letter
from datetime import datetime

# Function to extract BOM table from geodatabase and convert to XLSX
def extract_table_to_xlsx(gdb_path):
    try:
        arcpy.env.workspace = gdb_path
        tables = arcpy.ListTables('BOM')

        if tables:
            # Extract the parent folder name of the GDB file
            parent_folder = os.path.basename(os.path.dirname(gdb_path))

            # Create folder on desktop with today's date
            today_date = datetime.now().strftime("%Y-%m-%d")
            output_folder = os.path.join(os.path.expanduser('~'), 'Desktop', f'BOM_Pull_{today_date}')
            os.makedirs(output_folder, exist_ok=True)

            for table in tables:
                # Generate output XLSX file name based on the parent folder name
                output_xlsx = os.path.join(output_folder, f"{parent_folder}_{table}.xlsx")

                # Check if the file exists, if so, generate a unique name
                file_counter = 1
                while os.path.exists(output_xlsx):
                    output_xlsx = os.path.join(output_folder, f"{parent_folder}_{table}_{file_counter}.xlsx")
                    file_counter += 1

                # Convert table to XLSX format
                arcpy.TableToExcel_conversion(table, output_xlsx)
                print(f"Table '{table}' converted to '{output_xlsx}'")

                # Remove specific columns from the created XLSX files
                remove_columns(output_xlsx)
                adjust_column_width(output_xlsx)
        else:
            print(f"No 'BOM' table found in {gdb_path}")
    except arcpy.ExecuteError:
        print(arcpy.GetMessages())

# Function to remove specific columns from the created XLSX files
def remove_columns(file_path):
    try:
        columns_to_remove = ['V', 'U', 'T', 'S', 'R', 'Q', 'P', 'O', 'J', 'H', 'G', 'F', 'E', 'C', 'B', 'A']

        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in columns_to_remove:
                ws.delete_cols(ws[col + '1'].column)

        wb.save(file_path)
        print(f"Removed specific columns from '{os.path.basename(file_path)}'")
    except Exception as e:
        print(f"Error removing columns: {e}")

def adjust_column_width(file_path):
    try:
        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]

            columns_to_adjust = {
                'A': 120,
                'B': 170,
                'D': 719,
                'E': 90,
                'F': 90,
                'G': 90
            }  # Columns and their respective widths

            for col, width in columns_to_adjust.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width

                    # Check each cell in the column to determine the optimal width
                    max_length = 0
                    for cell in ws[col]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except TypeError:
                            pass

                    adjusted_width = (max_length + 2) * 1.2  # Adjust width (you can modify the multiplier for better fit)
                    if adjusted_width > width:
                        ws.column_dimensions[col].width = adjusted_width

        wb.save(file_path)
        print(f"Adjusted column width in '{os.path.basename(file_path)}'")
    except Exception as e:
        print(f"Error adjusting column width: {e}")

# Function to handle importing GDB list from Excel file
def import_gdb_list():
    global gdb_list
    file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            gdb_list = [row[0] for row in sheet.iter_rows(values_only=True)]
            update_listbox()
            print("GDB list imported successfully")
        except Exception as e:
            print(f"Error importing GDB list: {e}")

# Function to update the listbox with loaded GDB files
def update_listbox():
    listbox.delete(0, tk.END)
    for gdb_file in gdb_list:
        listbox.insert(tk.END, gdb_file)

# Function to start conversion process
def start_conversion():
    for gdb_file in gdb_list:
        extract_table_to_xlsx(gdb_file)
    messagebox.showinfo("BOM Tables Pulled", "BOM Pulled and converted to XLSX files!")

# Close the GUI
def close_gui():
    root.destroy()

# GUI setup
root = tk.Tk()
root.title("GDB Table Extractor")

gdb_list = []

# Import GDB list button
import_button = tk.Button(root, text="Import GDB List from Excel", command=import_gdb_list)
import_button.pack()

# Listbox to display loaded GDB files
listbox = tk.Listbox(root, selectmode=tk.MULTIPLE)
listbox.pack()

# Run button for starting conversion
run_button = tk.Button(root, text="Run", command=start_conversion)
run_button.pack()

# Close button
close_button = tk.Button(root, text="Close", command=close_gui)
close_button.pack()

root.mainloop()
