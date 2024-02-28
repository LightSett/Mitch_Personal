import os
import shutil
from openpyxl import load_workbook

def find_and_move_kmz(filepaths):
    for filepath in filepaths:
        st1_mapspace_folder = os.path.join(filepath, "st1_Mapspace")
        st1_exports_kmz_folder = os.path.join(filepath, "st1_Exports", "KMZ")
        
        # Check if the st1_Mapspace folder exists
        if not os.path.exists(st1_mapspace_folder):
            print(f"st1_Mapspace folder not found in {filepath}")
            continue
        
        # Create st1_Exports/KMZ folder if it doesn't exist
        os.makedirs(st1_exports_kmz_folder, exist_ok=True)
        
        # Search for UG_Structure_LayerToKML.kmz file in st1_Mapspace folder
        kmz_file = os.path.join(st1_mapspace_folder, "UG_Structure_LayerToKML.kmz")
        if os.path.exists(kmz_file):
            destination_path = os.path.join(st1_exports_kmz_folder, "UG_Structure_LayerToKML.kmz")
            shutil.move(kmz_file, destination_path)
            print(f"Moved UG_Structure_LayerToKML.kmz to {st1_exports_kmz_folder}.")
        else:
            print("UG_Structure_LayerToKML.kmz not found in st1_Mapspace folder.")

def get_filepaths_from_excel(excel_file):
    filepaths = []
    wb = load_workbook(excel_file)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        filepath = row[0]  # Assuming file paths are in the first column
        if filepath and os.path.isdir(filepath):
            filepaths.append(filepath)
    return filepaths

# Example usage:
excel_file_path = input("Enter the path to the Excel file: ")
if not os.path.isfile(excel_file_path):
    print("Invalid Excel file path.")
    exit()

filepaths_from_excel = get_filepaths_from_excel(excel_file_path)
if not filepaths_from_excel:
    print("No valid file paths found in Excel file.")
    exit()

find_and_move_kmz(filepaths_from_excel)
