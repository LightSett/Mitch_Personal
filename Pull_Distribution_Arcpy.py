import tkinter as tk
from tkinter import filedialog
import arcpy
import os
from openpyxl import Workbook, load_workbook

output_excel_path = ""  # Variable to store the path of the newly created Excel file

def create_data_sheet(gdb_path, output_excel):
    global output_excel_path
    try:
        arcpy.env.workspace = gdb_path
        feature_classes = ["Transmedia", "Equipment"]

        # Create Excel workbook
        excel_workbook = Workbook()

        for feature_class in feature_classes:
            try:
                # Create a new sheet for each feature class
                sheet = excel_workbook.create_sheet(title=feature_class)

                # Get field names
                fields = [field.name for field in arcpy.ListFields(feature_class)]
                sheet.append(fields)  # Write field names as headers

                # Retrieve attribute data
                with arcpy.da.SearchCursor(feature_class, fields) as cursor:
                    for row in cursor:
                        formatted_row = []
                        for item in row:
                            if isinstance(item, tuple):
                                formatted_coord = ', '.join(str(coord) for coord in item)
                                formatted_row.append(formatted_coord)
                            else:
                                formatted_row.append(item)
                        
                        sheet.append(formatted_row)

            except Exception as e:
                print(f"Error processing {feature_class}: {str(e)}")

        # Remove the default sheet created by openpyxl
        excel_workbook.remove(excel_workbook.active)

        # Save the Excel file
        with open(output_excel, "wb") as excel_file:
            excel_workbook.save(output_excel)
            print("Attribute tables exported to Excel successfully.")
            output_excel_path = output_excel
            return True

    except Exception as e:
        print(f"Error creating data sheet: {str(e)}")
        return False

def find_primary_distribution(gdb_path):
    try:
        excel_workbook = load_workbook(gdb_path)
        arcpy.env.workspace = gdb_path

        # Accessing the first sheet (index 0) as Transmedia
        transmedia_sheet = excel_workbook.worksheets[0]

        # Accessing the second sheet (index 1) as Equipment
        equipment_sheet = excel_workbook.worksheets[1]

        if equipment_sheet and transmedia_sheet:
            equipment_fc = "Equipment"
            equipment_field = "EQUIPMENT_TYPE"
            equipment_type = "PFP"

            equipment_objectid = None

            with arcpy.da.SearchCursor(equipment_fc, ["ObjectID"], f"{equipment_field} = '{equipment_type}'") as cursor:
                for row in cursor:
                    equipment_objectid = row[0]
                    break  # Assuming there's only one equipment with the specified type

            if equipment_objectid is not None:
                transmedia_fc = "Transmedia"
                transmedia_field = "From_Structure"

                primary_distribution_rows = []

                with arcpy.da.SearchCursor(transmedia_fc, [transmedia_field]) as cursor:
                    for row in cursor:
                        if row[0] == equipment_objectid:
                            primary_distribution_rows.append(row)

                return primary_distribution_rows

        else:
            print("Sheets 'Equipment' or 'Transmedia' not found in the Excel file.")
            return None

    except Exception as e:
        print(f"Error finding primary distribution: {str(e)}")
        return None




def copy_primary_distribution_to_excel(primary_distribution, output_excel):
    try:
        if primary_distribution:
            excel_workbook = Workbook()
            sheet = excel_workbook.active
            sheet.title = "Primary_Distribution"

            header_written = False
            for row in primary_distribution:
                if not header_written:
                    sheet.append([field.name for field in arcpy.ListFields("Transmedia")])
                    header_written = True
                sheet.append(row)

            excel_workbook.save(output_excel)
            print("Primary Distribution copied to Excel successfully.")
            return True

    except Exception as e:
        print(f"Error copying primary distribution to Excel: {str(e)}")
        return False

def run_create_data_sheet():
    gdb_path = create_data_sheet_entry.get()
    gdb_parent_folder = os.path.dirname(gdb_path)
    output_excel = os.path.join(gdb_parent_folder, f"{os.path.basename(gdb_parent_folder)}_Data_Sheet.xlsx")

    create_data_sheet(gdb_path, output_excel)

def run_find_and_copy_to_excel():
    gdb_path = find_and_copy_entry.get()
    gdb_parent_folder = os.path.dirname(gdb_path)
    output_excel = os.path.join(gdb_parent_folder, f"{os.path.basename(gdb_parent_folder)}_Primary_Distribution.xlsx")

    primary_distribution = find_primary_distribution(gdb_path)
    copy_primary_distribution_to_excel(primary_distribution, output_excel)

def browse_create_data_sheet_gdb():
    file_path = filedialog.askdirectory()
    if file_path:
        create_data_sheet_entry.delete(0, tk.END)
        create_data_sheet_entry.insert(0, file_path)

def browse_find_and_copy_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        find_and_copy_entry.delete(0, tk.END)
        find_and_copy_entry.insert(0, file_path)

def close_app():
    root.destroy()

root = tk.Tk()
root.title("ArcGIS Pro Script GUI")

create_data_sheet_entry = tk.Entry(root, width=50)
create_data_sheet_entry.pack(pady=5)

create_data_sheet_browse_button = tk.Button(root, text="Browse", command=browse_create_data_sheet_gdb)
create_data_sheet_browse_button.pack()

run_create_data_sheet_button = tk.Button(root, text="Create Data Sheet", command=run_create_data_sheet)
run_create_data_sheet_button.pack(pady=5)

find_and_copy_entry = tk.Entry(root, width=50)
find_and_copy_entry.pack(pady=5)

find_and_copy_browse_button = tk.Button(root, text="Browse", command=browse_find_and_copy_file)
find_and_copy_browse_button.pack()

run_find_and_copy_to_excel_button = tk.Button(root, text="Find and Copy to Excel", command=run_find_and_copy_to_excel)
run_find_and_copy_to_excel_button.pack(pady=5)

close_button = tk.Button(root, text="Close", command=close_app)
close_button.pack()

root.mainloop()
