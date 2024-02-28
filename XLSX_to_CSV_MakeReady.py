import openpyxl
import os
import csv

def update_column_with_filename(sheet, filename):
    # Set the header in column 3 (C) to "jobname"
    sheet.cell(row=1, column=3).value = "jobname"

    # Fill the rest of column 3 with the filename
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value:  # Check if the first cell in the row is populated
            sheet.cell(row=row, column=3).value = filename

    # Remove rows where the value in Column F is not equal to "pole"
    rows_to_delete = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=6).value != "pole":
            rows_to_delete.append(row)

    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)

def convert_xlsx_to_csv(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the filename without extension
    filename = os.path.splitext(os.path.basename(file_path))[0]

    # Apply modifications
    update_column_with_filename(sheet, filename)

    # Save as CSV
    csv_file_path = f"{filename}.csv"
    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)
        for row in sheet.iter_rows():
            csv_writer.writerow([cell.value for cell in row])

    return csv_file_path

def merge_all_csv_files(directory):
    csv_files = []
    # Iterate through all files in the directory
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            csv_file_path = convert_xlsx_to_csv(file_path)
            print(f"Converted '{filename}' to '{csv_file_path}'")
            csv_files.append(csv_file_path)

    # Merge all CSV files into one
    output_csv_file_path = os.path.join(directory, "Master_CSV.csv")
    with open(output_csv_file_path, 'w', newline='', encoding='utf-8') as output_csvfile:
        csv_writer = csv.writer(output_csvfile)
        for csv_file in csv_files:
            with open(csv_file, 'r', newline='', encoding='utf-8') as input_csvfile:
                csv_reader = csv.reader(input_csvfile)
                for row in csv_reader:
                    csv_writer.writerow(row)
            os.remove(csv_file)  # Remove individual CSV files after merging their content

    print("Master CSV file created.")

# Replace 'your_directory_path' with the path to your directory containing Excel files
directory_path = r'C:\Users\mmccarthy\Desktop\1101PA Make Ready Info\XLSX'
merge_all_csv_files(directory_path)
