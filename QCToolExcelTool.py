import openpyxl
import re

def manipulate_excel_data(input_file):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active

        # Create a new workbook
        new_wb = openpyxl.Workbook()

        # Iterate through each row in the original sheet
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            new_sheet = new_wb.create_sheet(title=f"Row_{row_index}")

            # Manipulate data in the new sheet
            new_row = []
            for cell in row:
                if cell.value:
                    new_row.extend(cell.value.split(';'))

            # Transpose data from one row to one column
            transposed_data = []
            for item in new_row:
                split_data = re.split(r',|/', item)
                transposed_data.append(split_data)

            # Split into columns by "," and "/"
            max_len = max(len(data) for data in transposed_data)
            for col_index in range(1, max_len + 1):
                for row_index, data in enumerate(transposed_data, start=1):
                    if col_index <= len(data):
                        new_sheet.cell(row=row_index, column=col_index, value=data[col_index - 1])

            # Remove brackets and quotation marks
            for row in new_sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value:
                        value = value.replace('[', '').replace(']', '').replace('"', '').replace("'", "")
                        cell.value = value

            # Rename sheet based on the value in A1
            new_sheet_name = new_sheet.cell(row=1, column=1).value
            new_sheet.title = re.sub(r'[^0-9.]', '', str(new_sheet_name))

            # Remove blank sheets
            empty_sheets = [sheet for sheet in new_wb.sheetnames if new_wb[sheet].cell(row=1, column=1).value is None]
            for sheet_name in empty_sheets:
                new_wb.remove(new_wb[sheet_name])

            # Resize columns to fit content
            for sheet in new_wb.sheetnames:
                current_sheet = new_wb[sheet]
                for col in range(1, current_sheet.max_column + 1):
                    max_length = 0
                    column = openpyxl.utils.get_column_letter(col)
                    for cell in current_sheet[openpyxl.utils.get_column_letter(col)]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except TypeError:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    current_sheet.column_dimensions[column].width = adjusted_width

        # Save the new workbook
        new_file_path = f"{input_file.replace('.xlsx', '')}_organized.xlsx"
        new_wb.save(new_file_path)
        print(f"New workbook '{new_file_path}' created successfully!")

    except Exception as e:
        print(f"Error manipulating data: {str(e)}")

# Replace 'your_file_path' with the path to your Excel file
file_path = r'replace/file/path/here'
manipulate_excel_data(file_path)
