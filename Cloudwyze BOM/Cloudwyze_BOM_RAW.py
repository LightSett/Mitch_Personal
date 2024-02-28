import arcpy
import os
from openpyxl import load_workbook

# Function to extract numbers separated by underscore from filepath
def extract_numbers(filepath):
    filename = os.path.basename(filepath)
    numbers = [part for part in filename.split('_') if part.isdigit()]
    return '_'.join(numbers[:2]) if len(numbers) >= 2 else None

# Path to the Excel file containing directory paths
excel_file_path = r"C:\Users\mmccarthy\Desktop\Automation Team\Lightsett\Cloudwyze BOM\Cloudwyze_AB_LLD_List.xlsx"

# Load the workbook
workbook = load_workbook(filename=excel_file_path)
sheet = workbook.active

# List to store directory paths
directory_paths = []

# Read directory paths from Excel file
for row in sheet.iter_rows(values_only=True):
    directory_paths.extend(row)

# Output directory (where newly created folders will be stored)
output_directory = os.path.dirname(os.path.abspath(__file__))
print(f"Output directory: {output_directory}")

arcpy.env.overwriteOutput = True

for input_directory in directory_paths:
    print(f"\nProcessing directory: {input_directory}")

    # Set the workspace to the current input directory
    arcpy.env.workspace = input_directory

    # List all feature classes in the input directory
    feature_classes = arcpy.ListFeatureClasses()

    if feature_classes:
        print("Found the following feature classes:")
        for fc in feature_classes:
            print(f"- {fc}")

    for fc in feature_classes:
        # Extract numbers from the filepath
        numbers = extract_numbers(fc)

        if numbers:
            # Create folder name based on extracted numbers
            folder_name = f"Cloudwyze_{numbers}_LLD"

            # Create the folder in the output directory
            new_folder_path = os.path.join(output_directory, folder_name)
            os.makedirs(new_folder_path, exist_ok=True)
            print(f"Created folder: {new_folder_path}")

            # Copy the feature class to the newly created folder
            arcpy.management.CopyFeatures(fc, os.path.join(new_folder_path, fc))
            print(f"Copied '{fc}' to '{new_folder_path}'")

print("\nProcess completed.")
