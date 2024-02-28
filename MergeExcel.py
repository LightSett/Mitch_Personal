import os
import openpyxl

# Specify the directory containing Excel files
directory = r'S:\Customers\23 V1Fiber\23.065.001 V1 Corning Lima\Scranton\Weekly Lima Report Data\PFP Locations'  # Change this to your specific location

# Create a new workbook to store the matching rows
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Iterate through each Excel file in the specified directory
for excel_file in os.listdir(directory):
    if excel_file.endswith('.xlsx'):
        file_path = os.path.join(directory, excel_file)

        # Open the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Iterate through each sheet in the Excel file
        for sheet in wb:
            # Iterate through each row in the sheet
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and "PFP" in str(cell):
                        # Copy the entire row to the new workbook
                        new_ws.append(row)

        # Close the Excel file
        wb.close()

# Save the new workbook with matching rows
new_file_path = r'S:\Customers\23 V1Fiber\23.065.001 V1 Corning Lima\Scranton\Weekly Lima Report Data\PFP Locations\combined.xlsx'  # Change this to the desired output path
new_wb.save(new_file_path)
new_wb.close()

print("Matching rows with 'PFP' in the entire workbook have been copied to a new workbook.")
